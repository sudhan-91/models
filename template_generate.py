"""
header_classifier.py
--------------------
Auto-classify a list of header column names into categories and produce
a color-coded Excel template matching the Multiproduct_template_v1 style.

Everything is fully dynamic — no hardcoded row numbers, freeze positions,
column widths, row heights, or font sizes.

Usage:
    from header_classifier import HeaderClassifier, HEADERS

    classifier = HeaderClassifier(HEADERS)
    path = classifier.generate_template("output.xlsx")
"""

import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────
# 1.  INPUT  – edit this list with your own headers
# ─────────────────────────────────────────────────
HEADERS = [
    "Product Name", "SP Number", "OPN", "HFGR", "Pre_M9",
    "M9_Release_Date", "Material_Status", "Product_Status",
    "Die Raster X", "Die Raster Y", "Die Thickness", "Wafer Diameter",
    "Process Node", "Product Technology", "ESD Data",
    "Wafer Test Temperature", "Final Test Temperature",
    "Product Package", "Product Package Type", "Body Length", "Body Width",
    "Body Thickness", "Net Weight", "Moulding Compound", "Leadframe",
    "Plating Surface", "Marking Format", "PIC No (Marking Pattern No)",
    "Small Packing Unit", "Large Packing Unit",
    "Wafer_Fab_Loc", "Wafer_Fab_Loc_Name", "Wafer_Test_Loc",
    "Wafer_Test_Loc_Name", "Assembly_Loc", "Assembly_Loc_Name",
    "Final_Test_Loc", "Final_Test_Loc_Name",
    "Halogen Free", "Completely Lead Free", "ROHS Compliance",
]


# ─────────────────────────────────────────────────
# 2.  MODULE-LEVEL CONSTANTS
# ─────────────────────────────────────────────────
CATEGORIES = {
    "Product attributes": {
        "keywords": [
            "product name", "sp number", "spnumber", "opn", "hfgr",
            "pre_m9", "pre m9", "m9_release", "m9 release",
            "material_status", "material status",
            "product_status", "product status",
            "die raster", "die_raster",
            "die thickness", "die_thickness",
            "wafer diameter", "wafer_diameter",
            "process node", "process_node",
            "product technology", "product_technology",
            "esd data", "esd_data",
            "wafer test temp", "wafer_test_temp",
            "final test temp", "final_test_temp",
            "ip number", "ip_number", "cpn", "gpn",
            "release date", "release_date",
        ],
        "cat_bg":   "FFD9E2F3",
        "cat_font": "FF000000",
        "hdr_bg":   "FF2E75B6",
        "hdr_font": "FFFFFFFF",
    },
    "Package / Packing attributes": {
        "keywords": [
            "package", "packing", "body length", "body_length",
            "body width", "body_width", "body thickness", "body_thickness",
            "net weight", "net_weight", "moulding", "molding",
            "leadframe", "lead frame", "plating", "marking",
            "pic no", "pic_no", "marking pattern",
            "small packing", "small_packing",
            "large packing", "large_packing",
            "packing unit", "packing_unit",
        ],
        "cat_bg":   "FFFFF2CB",
        "cat_font": "FF000000",
        "hdr_bg":   "FFBF9000",
        "hdr_font": "FFFFFFFF",
    },
    "Manufacturing location": {
        "keywords": [
            "wafer_fab", "wafer fab", "wafer_test_loc", "wafer test loc",
            "assembly_loc", "assembly loc", "assembly location",
            "final_test_loc", "final test loc",
            "_loc", "_loc_name", "loc name", "location",
            "fab loc", "fab_loc", "manufacturing",
        ],
        "cat_bg":   "FFE2EEDA",
        "cat_font": "FF000000",
        "hdr_bg":   "FF548135",
        "hdr_font": "FFFFFFFF",
    },
    "Sustainability": {
        "keywords": [
            "halogen", "lead free", "lead_free",
            "rohs", "compliance", "green", "eco",
            "environmental", "sustainab", "recyclable", "hazardous",
        ],
        "cat_bg":   "FFFBE4D5",
        "cat_font": "FF000000",
        "hdr_bg":   "FFB15D24",
        "hdr_font": "FFFFFFFF",
    },
}

UNKNOWN_CAT    = "Other"
UNKNOWN_COLORS = {
    "cat_bg": "FFD3D3D3", "cat_font": "FF000000",
    "hdr_bg": "FF808080", "hdr_font": "FFFFFFFF",
}

RED_FONT_KEYWORDS = [
    "die thickness", "die_thickness",
    "process node",  "process_node",
    "wafer test temp", "wafer_test_temp",
    "final test temp", "final_test_temp",
    "marking format",  "marking_format",
    "pic no", "pic_no", "marking pattern",
]

RED_FONT_COLOR   = "FFFF0000"
WHITE_FONT_COLOR = "FFFFFFFF"


# ─────────────────────────────────────────────────
# 3.  HEADER CLASSIFIER CLASS
# ─────────────────────────────────────────────────
class HeaderClassifier:
    """
    Classifies header column names into categories and writes
    a color-coded Excel template.

    Parameters
    ----------
    headers         : list of column name strings to classify.
    cat_header_row  : include category span row above column headers.
    font_name       : font family for all header cells.
    font_size       : font size (pts) for column header cells.
    cat_font_size   : font size (pts) for category header cells.
    row_height_cat  : row height (pts) for the category header row.
    row_height_hdr  : row height (pts) for the column header row.
    min_col_width   : minimum column width (chars).
    max_col_width   : maximum column width (chars).
    freeze_headers  : freeze all header rows on scroll.
    """

    def __init__(
        self,
        headers: list,
        *,
        cat_header_row: bool  = True,
        font_name:      str   = "Arial",
        font_size:      int   = 9,
        cat_font_size:  int   = 9,
        row_height_cat: float = 22,
        row_height_hdr: float = 40,
        min_col_width:  float = 12,
        max_col_width:  float = 45,
        freeze_headers: bool  = True,
    ):
        try:
            if not isinstance(headers, list) or not headers:
                raise ValueError("headers must be a non-empty list of column names.")

            self.headers        = headers
            self.cat_header_row = cat_header_row
            self.font_name      = font_name
            self.font_size      = font_size
            self.cat_font_size  = cat_font_size
            self.row_height_cat = row_height_cat
            self.row_height_hdr = row_height_hdr
            self.min_col_width  = min_col_width
            self.max_col_width  = max_col_width
            self.freeze_headers = freeze_headers

            logger.info("HeaderClassifier ready — %d columns.", len(headers))

        except Exception as e:
            logger.error("Init failed: %s", e)
            raise

    # ── Public entry point ────────────────────────────────────────────────────

    def generate_template(self, output_path: str = "header_classified.xlsx") -> str:
        """
        Classify headers → build styled Excel template → save → return path.
        """
        try:
            logger.info("Generating template → %s", output_path)
            classified = self._classify()
            wb         = self._build_workbook(classified)
            wb.save(output_path)
            logger.info(
                "Saved → %s  [%d cols, %d categories]",
                output_path, len(self.headers), len(classified),
            )
            return output_path

        except PermissionError as e:
            logger.error("Cannot write '%s' — file may be open: %s", output_path, e)
            raise
        except Exception as e:
            logger.error("Error in generate_template: %s", e)
            raise

    # ── Classifier ────────────────────────────────────────────────────────────

    def _classify(self) -> dict:
        """
        Group every header into a category preserving original order.
        Returns { category_name: [col, ...] }  (empty categories omitted).
        """
        try:
            logger.info("Classifying %d headers.", len(self.headers))

            col_pos = {c: i for i, c in enumerate(self.headers)}
            result  = {cat: [] for cat in CATEGORIES}
            result[UNKNOWN_CAT] = []

            for col in self.headers:
                bucket = self._best_category(col)
                result.setdefault(bucket, []).append(col)

            # Preserve original column order within each category
            for cat in result:
                result[cat].sort(key=lambda c: col_pos.get(c, 99_999))

            classified = {k: v for k, v in result.items() if v}
            logger.info(
                "Classification done: %s",
                {k: len(v) for k, v in classified.items()},
            )
            return classified

        except Exception as e:
            logger.error("Error in _classify: %s", e)
            raise

    def _best_category(self, col_name: str) -> str:
        """Score every category and return the best match for one column."""
        try:
            lower  = col_name.lower().strip()
            scores = {
                cat: sum(len(kw) for kw in cfg["keywords"] if kw in lower)
                for cat, cfg in CATEGORIES.items()
            }
            best = max(scores, key=lambda c: scores[c])
            return best if scores[best] > 0 else UNKNOWN_CAT

        except Exception as e:
            logger.error("Error in _best_category for '%s': %s", col_name, e)
            return UNKNOWN_CAT   # safe fallback → Other

    def _is_red_font(self, col_name: str) -> bool:
        """Return True if this column header should use red font."""
        try:
            lower = col_name.lower().strip()
            return any(kw in lower for kw in RED_FONT_KEYWORDS)

        except Exception as e:
            logger.error("Error in _is_red_font for '%s': %s", col_name, e)
            return False   # safe fallback → white font

    # ── Workbook builder ──────────────────────────────────────────────────────

    def _build_workbook(self, classified: dict) -> Workbook:
        """Construct and return the fully styled openpyxl Workbook."""
        try:
            # ── Dynamic row layout (nothing hardcoded) ─────────────────────
            # cat_header_row=True  → ROW_CAT=1, ROW_HDR=2, ROW_DATA=3
            # cat_header_row=False → ROW_CAT=None, ROW_HDR=1, ROW_DATA=2
            ROW_CAT  = 1 if self.cat_header_row else None
            ROW_HDR  = 2 if self.cat_header_row else 1
            ROW_DATA = ROW_HDR + 1       # first data row — always computed

            # col_name → 1-based Excel column index
            col_xi: dict = {}
            xi = 1
            for cat_cols in classified.values():
                for col in cat_cols:
                    col_xi[col] = xi
                    xi += 1

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            if ROW_CAT is not None:
                self._write_category_row(ws, classified, col_xi, ROW_CAT)

            self._write_header_row(ws, classified, col_xi, ROW_HDR)

            # Freeze panes computed from ROW_DATA — never a hardcoded string
            if self.freeze_headers:
                ws.freeze_panes = f"A{ROW_DATA}"
                logger.info("Freeze panes → A%d.", ROW_DATA)

            return wb

        except Exception as e:
            logger.error("Error in _build_workbook: %s", e)
            raise

    def _write_category_row(
        self, ws, classified: dict, col_xi: dict, row: int
    ) -> None:
        """Write merged category span headers onto *row*."""
        try:
            logger.info("Writing category row (row %d).", row)

            for cat, cat_cols in classified.items():
                start = col_xi[cat_cols[0]]
                end   = col_xi[cat_cols[-1]]
                cfg   = CATEGORIES.get(cat, UNKNOWN_COLORS)

                if start < end:
                    ws.merge_cells(
                        start_row=row, start_column=start,
                        end_row=row,   end_column=end,
                    )

                cell           = ws.cell(row=row, column=start, value=cat)
                cell.fill      = self._make_fill(cfg["cat_bg"])
                cell.font      = self._make_font(cfg["cat_font"], bold=True,
                                                 size=self.cat_font_size)
                cell.alignment = self._make_center_align()
                cell.border    = self._make_border()

            # Row height driven by self.row_height_cat — not hardcoded
            ws.row_dimensions[row].height = self.row_height_cat

        except Exception as e:
            logger.error("Error in _write_category_row (row %d): %s", row, e)
            raise

    def _write_header_row(
        self, ws, classified: dict, col_xi: dict, row: int
    ) -> None:
        """Write color-coded column headers onto *row*."""
        try:
            logger.info("Writing header row (row %d).", row)

            for cat, cat_cols in classified.items():
                cfg = CATEGORIES.get(cat, UNKNOWN_COLORS)
                for col in cat_cols:
                    idx  = col_xi[col]
                    cell = ws.cell(row=row, column=idx, value=col)

                    font_color     = RED_FONT_COLOR if self._is_red_font(col) else cfg["hdr_font"]
                    cell.fill      = self._make_fill(cfg["hdr_bg"])
                    cell.font      = self._make_font(font_color, bold=True,
                                                     size=self.font_size)
                    cell.alignment = self._make_center_align()
                    cell.border    = self._make_border()

                    # Width derived from header name length — not hardcoded
                    ws.column_dimensions[get_column_letter(idx)].width = (
                        self._col_width(col)
                    )

            # Row height driven by self.row_height_hdr — not hardcoded
            ws.row_dimensions[row].height = self.row_height_hdr

        except Exception as e:
            logger.error("Error in _write_header_row (row %d): %s", row, e)
            raise

    # ── Style builders ────────────────────────────────────────────────────────

    def _make_fill(self, hex_color: str) -> PatternFill:
        try:
            return PatternFill("solid", fgColor=hex_color)
        except Exception as e:
            logger.error("Error in _make_fill (hex=%s): %s", hex_color, e)
            return PatternFill("solid", fgColor="FFFFFFFF")   # fallback: white

    def _make_font(self, hex_color: str, bold: bool = True, size: int = None) -> Font:
        try:
            return Font(
                bold=bold,
                color=hex_color,
                name=self.font_name,
                size=size if size is not None else self.font_size,
            )
        except Exception as e:
            logger.error("Error in _make_font (hex=%s): %s", hex_color, e)
            return Font(bold=bold, name="Arial", size=9)   # fallback

    def _make_border(self) -> Border:
        try:
            side = Side(style="thin", color="FFD0D0D0")
            return Border(left=side, right=side, top=side, bottom=side)
        except Exception as e:
            logger.error("Error in _make_border: %s", e)
            return Border()   # fallback: no border

    def _make_center_align(self) -> Alignment:
        try:
            return Alignment(horizontal="center", vertical="center", wrap_text=True)
        except Exception as e:
            logger.error("Error in _make_center_align: %s", e)
            return Alignment()   # fallback

    def _col_width(self, col_name: str) -> float:
        """Width = header name length × scale, clamped to min/max."""
        try:
            raw = len(col_name) * 0.9 + 2
            return min(self.max_col_width, max(self.min_col_width, raw))
        except Exception as e:
            logger.error("Error in _col_width for '%s': %s", col_name, e)
            return self.min_col_width   # fallback


# ─────────────────────────────────────────────────
# 4.  MAIN
# ─────────────────────────────────────────────────
def main():
    try:
        print("=" * 60)
        print("Header Classifier & Template Generator")
        print("=" * 60)

        classifier = HeaderClassifier(HEADERS)
        classified = classifier._classify()

        print("\nClassification result:")
        total = 0
        for cat, cols in classified.items():
            print(f"\n  [{cat}]  ({len(cols)} columns)")
            for c in cols:
                flag = " ← RED FONT" if classifier._is_red_font(c) else ""
                print(f"    · {c}{flag}")
            total += len(cols)
        print(f"\nTotal: {total} columns across {len(classified)} categories\n")

        path = classifier.generate_template("header_classified.xlsx")
        print(f"Done! → {path}")

    except Exception as e:
        logger.error("Fatal error in main: %s", e)
        raise


if __name__ == "__main__":
    main()