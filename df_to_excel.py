"""
excel_exporter.py
-----------------
Usage:
    from excel_exporter import ExcelExporter

    exporter = ExcelExporter()
    path = exporter.generate_final_excel(df)
    path = exporter.generate_final_excel(df, "my_report.xlsx")
"""

from __future__ import annotations

import logging
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


class ExcelExporter:

    def __init__(self):
        try:
            self.categories = {
                "Product attributes": {
                    "keywords": [
                        "product name", "sp number", "spnumber", "opn", "hfgr",
                        "pre_m9", "pre m9", "m9_release", "m9 release",
                        "material_status", "material status",
                        "product_status", "product status",
                        "die raster", "die_raster",
                        "die thickness", "die_thickness",
                        "wafer diameter", "wafer_diameter",
                        "process node", "process_node",
                        "product technology", "product_technology",
                        "esd data", "esd_data",
                        "wafer test temp", "wafer_test_temp",
                        "final test temp", "final_test_temp",
                        "ip number", "ip_number", "cpn", "gpn",
                        "release date", "release_date",
                    ],
                    "cat_bg":   "FFD9E2F3",
                    "cat_font": "FF000000",
                    "hdr_bg":   "FF2E75B6",
                    "hdr_font": "FFFFFFFF",
                },
                "Package / Packing attributes": {
                    "keywords": [
                        "package", "packing", "body length", "body_length",
                        "body width", "body_width", "body thickness", "body_thickness",
                        "net weight", "net_weight", "moulding", "molding",
                        "leadframe", "lead frame", "plating", "marking",
                        "pic no", "pic_no", "marking pattern",
                        "small packing", "small_packing",
                        "large packing", "large_packing",
                        "packing unit", "packing_unit",
                    ],
                    "cat_bg":   "FFFFF2CB",
                    "cat_font": "FF000000",
                    "hdr_bg":   "FFBF9000",
                    "hdr_font": "FFFFFFFF",
                },
                "Manufacturing location": {
                    "keywords": [
                        "wafer_fab", "wafer fab", "wafer_test_loc", "wafer test loc",
                        "assembly_loc", "assembly loc", "assembly location",
                        "final_test_loc", "final test loc",
                        "_loc", "_loc_name", "loc name", "location",
                        "fab loc", "fab_loc", "manufacturing",
                    ],
                    "cat_bg":   "FFE2EEDA",
                    "cat_font": "FF000000",
                    "hdr_bg":   "FF548135",
                    "hdr_font": "FFFFFFFF",
                },
                "Sustainability": {
                    "keywords": [
                        "halogen", "lead free", "lead_free",
                        "rohs", "compliance", "green", "eco",
                        "environmental", "sustainab", "recyclable", "hazardous",
                    ],
                    "cat_bg":   "FFFBE4D5",
                    "cat_font": "FF000000",
                    "hdr_bg":   "FFB15D24",
                    "hdr_font": "FFFFFFFF",
                },
            }

            self.unknown_cat    = "Other"
            self.unknown_colors = {
                "cat_bg": "FFD3D3D3", "cat_font": "FF000000",
                "hdr_bg": "FF808080", "hdr_font": "FFFFFFFF",
            }

            self.red_font_keywords = [
                "die thickness", "die_thickness",
                "process node",  "process_node",
                "wafer test temp", "wafer_test_temp",
                "final test temp", "final_test_temp",
                "marking format",  "marking_format",
                "pic no", "pic_no", "marking pattern",
            ]

            self.row_bg          = ["FFFAFAFA", "FFFFFFFF"]
            self.row_height_cat  = 18
            self.row_height_hdr  = 38
            self.row_height_data = 15
            self.font_name       = "Arial"
            self.font_size       = 9
            self.min_col_width   = 10
            self.max_col_width   = 45

            logger.info("ExcelExporter initialised successfully.")

        except Exception as e:
            logger.error("Failed to initialise ExcelExporter: %s", e)
            raise

    # ── Public entry point ────────────────────────────────────────────────────

    def generate_final_excel(
        self,
        df: pd.DataFrame,
        output_path: str = "output.xlsx",
    ) -> str:
        """
        Receive a DataFrame → classify columns → write color-coded Excel.

        Parameters
        ----------
        df          : pandas DataFrame to export.
        output_path : Destination file path. Default: "output.xlsx".

        Returns
        -------
        str  Path of the saved Excel file.
        """
        try:
            logger.info("Starting Excel export → %s", output_path)

            if not isinstance(df, pd.DataFrame):
                raise TypeError(f"Expected a pandas DataFrame, got {type(df).__name__}.")
            if df.empty:
                raise ValueError("DataFrame is empty — nothing to export.")

            classified = self._classify(list(df.columns))
            wb         = self._build_workbook(df, classified)
            wb.save(output_path)

            logger.info(
                "Excel saved successfully → %s  [%d rows × %d cols]",
                output_path, len(df), len(df.columns),
            )
            return output_path

        except (TypeError, ValueError) as e:
            logger.error("Validation error in generate_final_excel: %s", e)
            raise
        except PermissionError as e:
            logger.error("Cannot write to '%s' — file may be open: %s", output_path, e)
            raise
        except Exception as e:
            logger.error("Unexpected error in generate_final_excel: %s", e)
            raise

    # ── Classifier ────────────────────────────────────────────────────────────

    def _classify(self, columns: list) -> dict:
        """Group every column into a category, preserving original order."""
        try:
            logger.info("Classifying %d columns into categories.", len(columns))

            col_pos = {c: i for i, c in enumerate(columns)}
            result  = {cat: [] for cat in self.categories}
            result[self.unknown_cat] = []

            for col in columns:
                bucket = self._best_category(col)
                result.setdefault(bucket, []).append(col)

            for cat in result:
                result[cat].sort(key=lambda c: col_pos.get(c, 99_999))

            classified = {k: v for k, v in result.items() if v}

            logger.info(
                "Classification complete: %s",
                {k: len(v) for k, v in classified.items()},
            )
            return classified

        except Exception as e:
            logger.error("Error in _classify: %s", e)
            raise

    def _best_category(self, col_name: str) -> str:
        """Score every category and return the best match for one column."""
        try:
            lower  = col_name.lower().strip()
            scores = {
                cat: sum(len(kw) for kw in cfg["keywords"] if kw in lower)
                for cat, cfg in self.categories.items()
            }
            best = max(scores, key=lambda c: scores[c])
            return best if scores[best] > 0 else self.unknown_cat

        except Exception as e:
            logger.error("Error in _best_category for column '%s': %s", col_name, e)
            return self.unknown_cat   # safe fallback — column goes to Other

    def _is_red_font(self, col_name: str) -> bool:
        """True if this column header should be rendered in red."""
        try:
            lower = col_name.lower().strip()
            return any(kw in lower for kw in self.red_font_keywords)

        except Exception as e:
            logger.error("Error in _is_red_font for column '%s': %s", col_name, e)
            return False   # safe fallback — default to non-red

    # ── Workbook builder ──────────────────────────────────────────────────────

    def _build_workbook(self, df: pd.DataFrame, classified: dict) -> Workbook:
        """Build and return the fully styled openpyxl Workbook."""
        try:
            logger.info("Building workbook structure.")

            # Dynamic row positions — derived, never hardcoded
            ROW_CAT  = 1
            ROW_HDR  = 2
            ROW_DATA = ROW_HDR + 1

            # Map: col_name → 1-based Excel column index
            col_xi: dict = {}
            xi = 1
            for cat_cols in classified.values():
                for col in cat_cols:
                    col_xi[col] = xi
                    xi += 1

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            self._write_category_row(ws, classified, col_xi, ROW_CAT)
            self._write_header_row(ws, df, classified, col_xi, ROW_HDR)
            self._write_data_rows(ws, df, col_xi, ROW_DATA)

            # Freeze panes: always computed from ROW_DATA, never hardcoded
            ws.freeze_panes = f"A{ROW_DATA}"
            logger.info("Freeze panes set at A%d.", ROW_DATA)

            return wb

        except Exception as e:
            logger.error("Error in _build_workbook: %s", e)
            raise

    def _write_category_row(
        self,
        ws,
        classified: dict,
        col_xi: dict,
        row: int,
    ) -> None:
        """Write the merged category span headers onto the given row."""
        try:
            logger.info("Writing category header row (row %d).", row)

            for cat, cat_cols in classified.items():
                start = col_xi[cat_cols[0]]
                end   = col_xi[cat_cols[-1]]
                cfg   = self.categories.get(cat, self.unknown_colors)

                if start < end:
                    ws.merge_cells(
                        start_row=row, start_column=start,
                        end_row=row,   end_column=end,
                    )

                cell           = ws.cell(row=row, column=start, value=cat)
                cell.fill      = self._make_fill(cfg["cat_bg"])
                cell.font      = self._make_font(cfg["cat_font"], bold=True)
                cell.alignment = self._make_center_align()
                cell.border    = self._make_border()

            ws.row_dimensions[row].height = self.row_height_cat

        except Exception as e:
            logger.error("Error in _write_category_row (row %d): %s", row, e)
            raise

    def _write_header_row(
        self,
        ws,
        df: pd.DataFrame,
        classified: dict,
        col_xi: dict,
        row: int,
    ) -> None:
        """Write color-coded column headers onto the given row."""
        try:
            logger.info("Writing column header row (row %d).", row)

            for cat, cat_cols in classified.items():
                cfg = self.categories.get(cat, self.unknown_colors)
                for col in cat_cols:
                    idx  = col_xi[col]
                    cell = ws.cell(row=row, column=idx, value=col)

                    font_color     = "FFFF0000" if self._is_red_font(col) else cfg["hdr_font"]
                    cell.fill      = self._make_fill(cfg["hdr_bg"])
                    cell.font      = self._make_font(font_color, bold=True)
                    cell.alignment = self._make_center_align()
                    cell.border    = self._make_border()

                    series = df[col] if col in df.columns else pd.Series(dtype=str)
                    ws.column_dimensions[get_column_letter(idx)].width = (
                        self._col_width(col, series)
                    )

            ws.row_dimensions[row].height = self.row_height_hdr

        except Exception as e:
            logger.error("Error in _write_header_row (row %d): %s", row, e)
            raise

    def _write_data_rows(
        self,
        ws,
        df: pd.DataFrame,
        col_xi: dict,
        start_row: int,
    ) -> None:
        """Write all data rows starting from start_row."""
        try:
            logger.info(
                "Writing %d data rows starting at row %d.", len(df), start_row
            )

            for row_i, (_, row) in enumerate(df.iterrows()):
                excel_row = start_row + row_i
                bg        = self.row_bg[row_i % len(self.row_bg)]

                for col, idx in col_xi.items():
                    try:
                        val = row.get(col, "")
                        if val is None or (isinstance(val, float) and np.isnan(val)):
                            val = ""
                    except Exception:
                        val = ""   # safe fallback for unreadable cell values

                    cell           = ws.cell(row=excel_row, column=idx, value=val)
                    cell.fill      = self._make_fill(bg)
                    cell.font      = self._make_font("FF000000", bold=False)
                    cell.alignment = self._make_left_align()
                    cell.border    = self._make_border()

                ws.row_dimensions[excel_row].height = self.row_height_data

        except Exception as e:
            logger.error("Error in _write_data_rows: %s", e)
            raise

    # ── Style builders ────────────────────────────────────────────────────────

    def _make_fill(self, hex_color: str) -> PatternFill:
        try:
            return PatternFill("solid", fgColor=hex_color)
        except Exception as e:
            logger.error("Error in _make_fill (hex_color=%s): %s", hex_color, e)
            return PatternFill("solid", fgColor="FFFFFFFF")   # fallback: white

    def _make_font(self, hex_color: str, bold: bool = True) -> Font:
        try:
            return Font(
                bold=bold,
                color=hex_color,
                name=self.font_name,
                size=self.font_size,
            )
        except Exception as e:
            logger.error("Error in _make_font (hex_color=%s): %s", hex_color, e)
            return Font(bold=bold, name="Arial", size=9)   # fallback: plain font

    def _make_border(self) -> Border:
        try:
            side = Side(style="thin", color="FFD0D0D0")
            return Border(left=side, right=side, top=side, bottom=side)
        except Exception as e:
            logger.error("Error in _make_border: %s", e)
            return Border()   # fallback: no border

    def _make_center_align(self) -> Alignment:
        try:
            return Alignment(horizontal="center", vertical="center", wrap_text=True)
        except Exception as e:
            logger.error("Error in _make_center_align: %s", e)
            return Alignment()   # fallback: default alignment

    def _make_left_align(self) -> Alignment:
        try:
            return Alignment(horizontal="left", vertical="center", wrap_text=False)
        except Exception as e:
            logger.error("Error in _make_left_align: %s", e)
            return Alignment()   # fallback: default alignment

    def _col_width(self, col_name: str, series: pd.Series) -> float:
        try:
            max_data = int(series.astype(str).str.len().max()) if len(series) > 0 else 0
            raw      = max(len(col_name), max_data) * 0.88 + 2
            return min(self.max_col_width, max(self.min_col_width, raw))
        except Exception as e:
            logger.error("Error in _col_width for column '%s': %s", col_name, e)
            return self.min_col_width   # fallback: minimum width


export = ExcelExporter()
df = pd.read_excel("/home/sudhan/Downloads/product_data.xlsx")
export.generate_final_excel(df)